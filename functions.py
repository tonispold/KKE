import os
import shutil
import time

import openpyxl
from openpyxl.styles import PatternFill
from zipfile import ZipFile


def cell_string(wb, sheet_name, cell_name):
    sheet = wb[sheet_name]
    cell = sheet[cell_name]
    return cell.value


def is_formula(cell_name):
    if str(cell_name) is None:
        return False
    if str(cell_name)[0] == '=':
        return True


def cell_change_colour(wb, sheet_name, cell_name, colour):
    sheet = wb[sheet_name]
    cell = sheet[cell_name]
    cell.fill = PatternFill(start_color=colour, end_color=colour, fill_type="solid")


def cell_write(wb, sheet_name, cell_name, value):
    sheet = wb[sheet_name]
    cell = sheet[cell_name]
    cell.value = value


def cell_answer(file_path, sheet_name, cell_name):
    wb2 = openpyxl.load_workbook(file_path, data_only=True)
    sheet = wb2[sheet_name]
    cell = sheet[cell_name]
    return cell.value


def delete_excel_cell_formating(wb, sheet_name, cell_name):
    sheet = wb[sheet_name]
    cell = sheet[cell_name]
    cell.style = 'Normal'


def sum_count(student_file, wb):
    lists = "sum, count"

    good = []
    bad = []
    empty = []

    not_formula = []
    formula_error = []
    wrong_answer = []

    count_if = dict(G29=4, G30=5, G31=8, G32=6, G33=9, G42=2, G43=2, G44=2, G45=9)

    sum_if = dict(G36=105, G37=164, G38=156, G39=511, G47=25, G48=75, G49=309, G52=386)

    for key, value in count_if.items():
        formula = cell_string(wb, lists, key)
        if check_if_cell_empty(cell_answer(student_file, lists, key)):
            empty.append(key)
            bad.append(key)
            continue
        else:
            if is_formula(formula):
                if formula.find('COUNTIF') != -1 or formula.find('COUNTIFS') != -1:
                    delete_excel_cell_formating(wb, lists, key)
                    if cell_answer(student_file, lists, key) == value:
                        good.append(key)
                    else:
                        bad.append(key)
                        wrong_answer.append(key)
                else:
                    bad.append(key)
                    formula_error.append(key)
            else:
                bad.append(key)
                not_formula.append(key)

    for key, value in sum_if.items():

        formula = cell_string(wb, lists, key)
        if check_if_cell_empty(cell_answer(student_file, lists, key)):
            empty.append(key)
            bad.append(key)
            continue
        else:
            if is_formula(formula):
                if formula.find('SUMIF') != -1 or formula.find('SUMIFS') != -1:
                    delete_excel_cell_formating(wb, lists, key)
                    if cell_answer(student_file, lists, key) == value:
                        good.append(key)
                    else:
                        bad.append(key)
                        wrong_answer.append(key)
                else:
                    bad.append(key)
                    formula_error.append(key)
            else:
                bad.append(key)
                not_formula.append(key)

    for i in good:
        cell_change_colour(wb, lists, i.replace("G", "F"), "33FF33")
        cell_change_colour(wb, lists, i, "33FF33")

    for i in bad:
        cell_change_colour(wb, lists, i.replace("G", "F"), "FF6666")
        cell_change_colour(wb, lists, i, "FF6666")

    for i in not_formula:
        cell_write(wb, lists, i.replace("G", "H"), "Not a function")

    for i in formula_error:
        cell_write(wb, lists, i.replace("G", "H"), "Wrong function used")

    for i in wrong_answer:
        cell_write(wb, lists, i.replace("G", "H"), "Wrong answer")

    for i in empty:
        cell_write(wb, lists, i.replace("G", "H"), "Cell is empty")


def check_if_cell_empty(value):
    if value is None:
        return True
    else:
        return False


def text_functions(student_file, wb):
    lists = "text functions"

    good = []
    bad = []
    empty = []

    not_formula = []
    wrong_answer = []
    formula_error = []

    fields = dict(B2="MA", B3="CA", B4="CA", B5="AZ", B6="TX", B10="1BB2", B11="1PT", B12="1Z", B13="D", B14="1V24C",
                  B15="1AA", B16="1ZFD3", C10="AC12", C11="AB34", C12="CD8", C13="PO65S3", C14="BV45", C15="DS96S",
                  C16="CD90")
    if_fields = dict(B26=1300, D30="Pass", D31="Pass", D32="Fail", D33="Pass", D34="Pass", D35="Pass", D36="Fail")

    for key, value in fields.items():

        formula = cell_string(wb, lists, key)
        if check_if_cell_empty(cell_answer(student_file, lists, key)):
            empty.append(key)
            bad.append(key)
            continue
        else:
            if is_formula(formula):
                delete_excel_cell_formating(wb, lists, key)
                if str(cell_answer(student_file, lists, key)) == str(value):
                    good.append(key)
                else:
                    bad.append(key)
                    wrong_answer.append(key)
            else:
                bad.append(key)
                not_formula.append(key)

    for key, value in if_fields.items():
        formula = cell_string(wb, lists, key)
        if check_if_cell_empty(cell_answer(student_file, lists, key)):
            empty.append(key)
            bad.append(key)
            continue
        else:
            if is_formula(formula):
                if formula.find('IF') != -1 or formula.find('IFS') != -1:
                    delete_excel_cell_formating(wb, lists, key)
                    if cell_answer(student_file, lists, key) == value:
                        good.append(key)
                    else:
                        bad.append(key)
                        wrong_answer.append(key)
                else:
                    bad.append(key)
                    formula_error.append(key)
            else:
                bad.append(key)
                not_formula.append(key)

    for i in good:
        cell_change_colour(wb, lists, i, "33FF33")

    for i in bad:
        cell_change_colour(wb, lists, i, "FF6666")

    for i in not_formula:
        if str(i).find("B") != -1:
            cell_write(wb, lists, i.replace("B", "E"), "Not a function")
            cell_change_colour(wb, lists, i.replace("B", "E"), "FDDA0D")
            cell_change_colour(wb, lists, i.replace("B", "F"), "FDDA0D")
        if str(i).find("C") != -1:
            cell_write(wb, lists, i.replace("C", "F"), "Not a function")
            cell_change_colour(wb, lists, i.replace("C", "F"), "FDDA0D")
        if str(i).find("D") != -1:
            cell_write(wb, lists, i.replace("D", "G"), "Not a function")
            cell_change_colour(wb, lists, i.replace("D", "G"), "FDDA0D")
            cell_change_colour(wb, lists, i.replace("D", "H"), "FDDA0D")

    for i in formula_error:
        if str(i).find("B") != -1:
            cell_write(wb, lists, i.replace("B", "E"), "Wrong function used")
            cell_change_colour(wb, lists, i.replace("B", "E"), "FDDA0D")
            cell_change_colour(wb, lists, i.replace("B", "F"), "FDDA0D")
        if str(i).find("C") != -1:
            cell_write(wb, lists, i.replace("C", "F"), "Wrong function used")
            cell_change_colour(wb, lists, i.replace("C", "F"), "FDDA0D")
        if str(i).find("D") != -1:
            cell_write(wb, lists, i.replace("D", "G"), "Wrong function used")
            cell_change_colour(wb, lists, i.replace("D", "G"), "FDDA0D")
            cell_change_colour(wb, lists, i.replace("D", "H"), "FDDA0D")

    for i in wrong_answer:
        if str(i).find("B") != -1:
            cell_write(wb, lists, i.replace("B", "E"), "Wrong answer")
            cell_change_colour(wb, lists, i.replace("B", "E"), "FDDA0D")
            cell_change_colour(wb, lists, i.replace("B", "F"), "FDDA0D")
        if str(i).find("C") != -1:
            cell_write(wb, lists, i.replace("C", "F"), "Wrong answer")
            cell_change_colour(wb, lists, i.replace("C", "F"), "FDDA0D")
        if str(i).find("D") != -1:
            cell_write(wb, lists, i.replace("D", "G"), "Wrong answer")
            cell_change_colour(wb, lists, i.replace("D", "G"), "FDDA0D")
            cell_change_colour(wb, lists, i.replace("D", "H"), "FDDA0D")

    for i in empty:
        if str(i).find("B") != -1:
            cell_write(wb, lists, i.replace("B", "E"), "Cell is empty")
            cell_change_colour(wb, lists, i.replace("B", "E"), "FDDA0D")
            cell_change_colour(wb, lists, i.replace("B", "F"), "FDDA0D")
        if str(i).find("C") != -1:
            cell_write(wb, lists, i.replace("C", "F"), "Cell is empty")
            cell_change_colour(wb, lists, i.replace("C", "F"), "FDDA0D")
        if str(i).find("D") != -1:
            cell_write(wb, lists, i.replace("D", "G"), "Cell is empty")
            cell_change_colour(wb, lists, i.replace("D", "G"), "FDDA0D")
            cell_change_colour(wb, lists, i.replace("D", "H"), "FDDA0D")


def date_functions(student_file, wb):
    lists = "Date functions"

    good = []
    bad = []
    empty = []

    not_formula = []
    formula_error = []
    wrong_answer = []

    dates = dict(E3="2020-03-01 00:00:00", E4="2019-06-03 00:00:00", E6=8, I3="2020-04-30 00:00:00",
                 I4="2019-03-03 00:00:00")
    week = dict(F3=1, F4=2)
    end = dict(G3="2020-04-30 00:00:00", G4="2019-07-31 00:00:00")
    func = dict(B8="=NOW()", D8="=HOUR(B8)", G8="=MINUTE(B8)")

    for key, value in dates.items():

        formula = cell_string(wb, lists, key)
        if check_if_cell_empty(cell_answer(student_file, lists, key)):
            empty.append(key)
            bad.append(key)
            continue
        else:
            if is_formula(formula):
                if formula.find('DATE') != -1 or formula.find('DATES') != -1:
                    delete_excel_cell_formating(wb, lists, key)
                    if str(cell_answer(student_file, lists, key)) == str(value):
                        good.append(key)
                    else:
                        bad.append(key)
                        wrong_answer.append(key)
                else:
                    bad.append(key)
                    formula_error.append(key)
            else:
                bad.append(key)
                not_formula.append(key)

    for key, value in week.items():

        formula = cell_string(wb, lists, key)
        if check_if_cell_empty(cell_answer(student_file, lists, key)):
            empty.append(key)
            bad.append(key)
            continue
        else:
            if is_formula(formula):
                if formula.find('WEEKDAY') != -1 or formula.find('WEEKDAYS') != -1:
                    delete_excel_cell_formating(wb, lists, key)
                    if cell_answer(student_file, lists, key) == value:
                        good.append(key)
                    else:
                        bad.append(key)
                        wrong_answer.append(key)
                else:
                    bad.append(key)
                    formula_error.append(key)
            else:
                bad.append(key)
                not_formula.append(key)

    for key, value in end.items():

        formula = cell_string(wb, lists, key)
        if check_if_cell_empty(cell_answer(student_file, lists, key)):
            empty.append(key)
            bad.append(key)
            continue
        else:
            if is_formula(formula):
                if formula.find('EOMONTH') != -1 or formula.find('EOMONTHS') != -1:
                    delete_excel_cell_formating(wb, lists, key)
                    if str(cell_answer(student_file, lists, key)) == str(value):
                        good.append(key)
                    else:
                        bad.append(key)
                        wrong_answer.append(key)
                else:
                    bad.append(key)
                    formula_error.append(key)
            else:
                bad.append(key)
                not_formula.append(key)

    for key, value in func.items():

        formula = cell_string(wb, lists, key)
        if check_if_cell_empty(cell_answer(student_file, lists, key)):
            empty.append(key)
            bad.append(key)
            continue
        else:
            if is_formula(formula):
                delete_excel_cell_formating(wb, lists, key)
                if formula.find(value) != -1:
                    good.append(key)
                else:
                    bad.append(key)
                    wrong_answer.append(key)
            else:
                bad.append(key)
                not_formula.append(key)

    formula_test = cell_string(wb, lists, "E5")
    if check_if_cell_empty(cell_answer(student_file, lists, "E5")):
        empty.append("E5")
        bad.append("E5")
    else:
        if is_formula(formula_test):
            delete_excel_cell_formating(wb, lists, "E5")
            if cell_answer(student_file, lists, "E5") == 272:
                good.append("E5")
            else:
                bad.append("E5")
                wrong_answer.append("E5")
        else:
            bad.append("E5")
            not_formula.append("E5")

    for i in good:
        cell_change_colour(wb, lists, i, "33FF33")

    for i in bad:
        cell_change_colour(wb, lists, i, "FF6666")

    for i in not_formula:
        vals = str(i[0]) + str(int(i[1:]) + 7)
        cell_write(wb, lists, vals, "Not a function")
        cell_change_colour(wb, lists, vals, "FDDA0D")

    for i in formula_error:
        vals = str(i[0]) + str(int(i[1:]) + 7)
        cell_write(wb, lists, vals, "Wrong function used")
        cell_change_colour(wb, lists, vals, "FDDA0D")

    for i in wrong_answer:
        vals = str(i[0]) + str(int(i[1:]) + 7)
        cell_write(wb, lists, vals, "Wrong answer")
        cell_change_colour(wb, lists, vals, "FDDA0D")

    for i in empty:
        vals = str(i[0]) + str(int(i[1:]) + 7)
        cell_write(wb, lists, vals, "Cell is empty")
        cell_change_colour(wb, lists, vals, "FDDA0D")


def logical_functions(student_file, wb):
    lists = "Logical functions"

    good_if = []
    good_else = []
    bad_if = []
    bad_else = []
    empty = []
    empty_else = []

    not_formula_if = []
    formula_error_if = []
    wrong_answer_if = []
    not_formula_else = []
    formula_error_else = []
    wrong_answer_else = []

    if_condition = dict(D2=0.0325, D3='On target', D4='On target', D5=0.365, D6=0.54, D7=0.13, D8='On target', D9=1.27)
    if_and_condition = dict(D19=133, D20=250, D21=700, D22=267, D23=300, D24=300)
    if_average_condition = dict(E19="decommissioned", E20="decommissioned", E21="ok", E22="decommissioned", E23="ok",
                                E24="ok")

    for key, value in if_condition.items():
        formula = cell_string(wb, lists, key)
        if check_if_cell_empty(cell_answer(student_file, lists, key)):
            empty.append(key)
            bad_if.append(key)
            continue
        else:
            if is_formula(formula):
                if formula.find('IF') != -1:
                    delete_excel_cell_formating(wb, lists, key)
                    if cell_answer(student_file, lists, key) == value:
                        good_if.append(key)
                    else:
                        bad_if.append(key)
                        wrong_answer_if.append(key)
                else:
                    bad_if.append(key)
                    formula_error_if.append(key)
            else:
                bad_if.append(key)
                not_formula_if.append(key)

    for key, value in if_and_condition.items():
        formula = cell_string(wb, lists, key)
        if check_if_cell_empty(cell_answer(student_file, lists, key)):
            empty_else.append(key)
            bad_else.append(key)
            continue
        else:
            if is_formula(formula):
                if formula.find('IF') != -1 and formula.find('AND') != -1:
                    delete_excel_cell_formating(wb, lists, key)
                    if round(cell_answer(student_file, lists, key)) == value:
                        good_else.append(key)
                    else:
                        bad_else.append(key)
                        wrong_answer_else.append(key)
                else:
                    bad_else.append(key)
                    formula_error_else.append(key)
            else:
                bad_else.append(key)
                not_formula_else.append(key)

    for key, value in if_average_condition.items():
        formula = cell_string(wb, lists, key)
        if check_if_cell_empty(cell_answer(student_file, lists, key)):
            empty_else.append(key)
            bad_else.append(key)
            continue
        else:
            if is_formula(formula):
                if formula.find('IF') != -1 and formula.find('AVERAGE') != -1:
                    delete_excel_cell_formating(wb, lists, key)
                    if cell_answer(student_file, lists, key) == value:
                        good_else.append(key)
                    else:
                        bad_else.append(key)
                        wrong_answer_else.append(key)
                else:
                    bad_else.append(key)
                    formula_error_else.append(key)
            else:
                bad_else.append(key)
                not_formula_else.append(key)

    for i in good_if:
        cell_change_colour(wb, lists, i, "33FF33")

    for i in bad_if:
        cell_change_colour(wb, lists, i, "FF6666")

    for i in not_formula_if:
        cell_write(wb, lists, i.replace("D", "E"), "Not a function")
        cell_change_colour(wb, lists, i.replace("D", "E"), "FDDA0D")

    for i in formula_error_if:
        cell_write(wb, lists, i.replace("D", "E"), "Wrong function used")
        cell_change_colour(wb, lists, i.replace("D", "E"), "FDDA0D")

    for i in wrong_answer_if:
        cell_write(wb, lists, i.replace("D", "E"), "Wrong answer")
        cell_change_colour(wb, lists, i.replace("D", "E"), "FDDA0D")

    for i in empty:
        cell_write(wb, lists, i.replace("D", "E"), "Cell is empty")
        cell_change_colour(wb, lists, i.replace("D", "E"), "FDDA0D")

    for i in good_else:
        cell_change_colour(wb, lists, i, "33FF33")

    for i in bad_else:
        cell_change_colour(wb, lists, i, "FF6666")

    for i in not_formula_else:
        if i.find("D") != -1:
            cell_write(wb, lists, i.replace("D", "F"), "Not a function")
            cell_change_colour(wb, lists, i.replace("D", "F"), "FDDA0D")
        if i.find("E") != -1:
            cell_write(wb, lists, i.replace("E", "G"), "Not a function")
            cell_change_colour(wb, lists, i.replace("E", "G"), "FDDA0D")

    for i in formula_error_else:
        if i.find("D") != -1:
            cell_write(wb, lists, i.replace("D", "F"), "Wrong function used")
            cell_change_colour(wb, lists, i.replace("D", "F"), "FDDA0D")
        if i.find("E") != -1:
            cell_write(wb, lists, i.replace("E", "G"), "Wrong function used")
            cell_change_colour(wb, lists, i.replace("E", "G"), "FDDA0D")

    for i in wrong_answer_else:
        if i.find("D") != -1:
            cell_write(wb, lists, i.replace("D", "F"), "Wrong answer")
            cell_change_colour(wb, lists, i.replace("D", "F"), "FDDA0D")
        if i.find("E") != -1:
            cell_write(wb, lists, i.replace("E", "G"), "Wrong answer")
            cell_change_colour(wb, lists, i.replace("E", "G"), "FDDA0D")

    for i in empty_else:
        if i.find("D") != -1:
            cell_write(wb, lists, i.replace("D", "F"), "Cell is empty")
            cell_change_colour(wb, lists, i.replace("D", "F"), "FDDA0D")
        if i.find("E") != -1:
            cell_write(wb, lists, i.replace("E", "G"), "Cell is empty")
            cell_change_colour(wb, lists, i.replace("E", "G"), "FDDA0D")


def lookup_functions(student_file, wb):
    lists = "lookup functions"
    months = ["Jan", "Feb", "Mar", "Apr", "May", "Jun"]

    good = []
    bad = []
    empty = []

    not_formula = []
    formula_error = []
    wrong_answer = []

    look_up = dict(F6=200, F7=25, F8=150, F9=75, F10=750, F11=75, F12=300, F13=150, F14=25, F15=150, F16=150)

    for key, value in look_up.items():
        formula = cell_string(wb, lists, key)
        if check_if_cell_empty(cell_answer(student_file, lists, key)):
            empty.append(key)
            bad.append(key)
            continue
        else:
            if is_formula(formula):
                if formula.find('VLOOKUP') != -1:
                    delete_excel_cell_formating(wb, lists, key)
                    if cell_answer(student_file, lists, key) == value:
                        good.append(key)
                    else:
                        bad.append(key)
                        wrong_answer.append(key)
                else:
                    bad.append(key)
                    formula_error.append(key)
            else:
                bad.append(key)
                not_formula.append(key)

    jaanuar = dict(Week1=[1, 1, 272], Week2=[1, 2, 112], Week3=[1, 3, 282], Week4=[1, 4, 114])
    veebruar = dict(Week1=[2, 1, 251], Week2=[2, 2, 363], Week3=[2, 3, 59], Week4=[2, 4, 421])
    march = dict(Week1=[3, 1, 339], Week2=[3, 2, 162], Week3=[3, 3, 409], Week4=[3, 4, 438])
    april = dict(Week1=[4, 1, 412], Week2=[4, 2, 269], Week3=[4, 3, 215], Week4=[4, 4, 391])
    mai = dict(Week1=[5, 1, 16], Week2=[5, 2, 358], Week3=[5, 3, 342], Week4=[5, 4, 110])
    june = dict(Week1=[6, 1, 137], Week2=[6, 2, 334], Week3=[6, 3, 429], Week4=[6, 4, 181])

    arr = [jaanuar, veebruar, march, april, mai, june]
    excel_cell_month_value = str(cell_string(wb, lists, "C48"))
    month = months.index(excel_cell_month_value)
    week = cell_string(wb, lists, "C49")
    answers = dict(C51=[arr[month][week][0], "MATCH"], C52=[arr[month][week][1], "MATCH"],
                   C54=[arr[month][week][2], "INDEX"], C56=[arr[month][week][2], "INDIRECT"],
                   C58=[arr[month][week][2], "OFFSET"])

    for key, value in answers.items():
        formula = cell_string(wb, lists, key)
        if check_if_cell_empty(cell_answer(student_file, lists, key)):
            empty.append(key)
            bad.append(key)
            continue
        else:
            if is_formula(formula):
                if formula.find(value[1]) != -1:
                    delete_excel_cell_formating(wb, lists, key)
                    if cell_answer(student_file, lists, key) == value[0]:
                        good.append(key)
                    else:
                        bad.append(key)
                        wrong_answer.append(key)
                else:
                    bad.append(key)
                    formula_error.append(key)
            else:
                bad.append(key)
                not_formula.append(key)

    categ = ["OVH (overheads)", "MAT (material)", "OGS (other goods/services)", "SAL (salaries)", "DEP (depreciation)"]

    categ2 = [["ovhCategory", 1177], ["matCategory", 761], ["ogsCategory", 1385], ["salCategory", 2013],
              ["depCategory", 1003]]
    value_of = str(cell_string(wb, lists, "D31"))
    cat = categ.index(value_of)

    formula = cell_string(wb, lists, "D33")
    formula2 = cell_string(wb, lists, "D35")
    if check_if_cell_empty(cell_answer(student_file, lists, "D33")) or check_if_cell_empty(
            cell_answer(student_file, lists, "D35")):
        empty.append("D33")
        bad.append("D33")

        empty.append("D35")
        bad.append("D35")
    else:
        if is_formula(formula) and is_formula(formula2):
            if formula.find("CONCATENATE") != -1 and formula2.find("SUM") != -1:
                delete_excel_cell_formating(wb, lists, "D33")
                delete_excel_cell_formating(wb, lists, "D35")
                if str(cell_answer(student_file, lists, "D33")) == str(categ2[cat][0]) and str(
                        cell_answer(student_file, lists, "D35")) == str(categ2[cat][1]):
                    good.append("D33")
                    good.append("D35")
                else:
                    bad.append("D33")
                    bad.append("D35")
                    wrong_answer.append("D33")
                    wrong_answer.append("D35")
            else:
                bad.append("D33")
                bad.append("D35")
                formula_error.append("D33")
                formula_error.append("D35")
        else:
            bad.append("D33")
            bad.append("D35")
            not_formula.append("D33")
            not_formula.append("D35")

    for i in good:
        cell_change_colour(wb, lists, i, "33FF33")

    for i in bad:
        cell_change_colour(wb, lists, i, "FF6666")

    for i in not_formula:
        if i.find("F") != -1:
            cell_write(wb, lists, i.replace("F", "G"), "Not a function")
            cell_change_colour(wb, lists, i.replace("F", "G"), "FDDA0D")
        if i.find("C") != -1:
            cell_write(wb, lists, i.replace("C", "F"), "Not a function")
            cell_change_colour(wb, lists, i.replace("C", "F"), "FDDA0D")
        if i.find("D") != -1:
            cell_write(wb, lists, i.replace("D", "I"), "Not a function")
            cell_change_colour(wb, lists, i.replace("D", "I"), "FDDA0D")

    for i in formula_error:
        if i.find("F") != -1:
            cell_write(wb, lists, i.replace("F", "G"), "Wrong function used")
            cell_change_colour(wb, lists, i.replace("F", "G"), "FDDA0D")
        if i.find("C") != -1:
            cell_write(wb, lists, i.replace("C", "F"), "Wrong function used")
            cell_change_colour(wb, lists, i.replace("C", "F"), "FDDA0D")
        if i.find("D") != -1:
            cell_write(wb, lists, i.replace("D", "I"), "Wrong function used")
            cell_change_colour(wb, lists, i.replace("D", "I"), "FDDA0D")

    for i in wrong_answer:
        if i.find("F") != -1:
            cell_write(wb, lists, i.replace("F", "G"), "Wrong answer")
            cell_change_colour(wb, lists, i.replace("F", "G"), "FDDA0D")
        if i.find("C") != -1:
            cell_write(wb, lists, i.replace("C", "F"), "Wrong answer")
            cell_change_colour(wb, lists, i.replace("C", "F"), "FDDA0D")
        if i.find("D") != -1:
            cell_write(wb, lists, i.replace("D", "I"), "Wrong answer")
            cell_change_colour(wb, lists, i.replace("D", "I"), "FDDA0D")

    for i in empty:
        if i.find("F") != -1:
            cell_write(wb, lists, i.replace("F", "G"), "Cell is empty")
            cell_change_colour(wb, lists, i.replace("F", "G"), "FDDA0D")
        if i.find("C") != -1:
            cell_write(wb, lists, i.replace("C", "F"), "Cell is empty")
            cell_change_colour(wb, lists, i.replace("C", "F"), "FDDA0D")
        if i.find("D") != -1:
            cell_write(wb, lists, i.replace("D", "I"), "Cell is empty")
            cell_change_colour(wb, lists, i.replace("D", "I"), "FDDA0D")


def list_from_txt(file):
    with open(file, "r") as f:
        lines = f.readlines()
    return lines


def copy_file(file):
    shutil.copy(file, file.replace(".xlsx", "_copy.xlsx"))


def create_zip(name):
    zip_obj = ZipFile(name + '.zip', 'w')
    zip_obj.close()


def add_file_to_zip_without_directory(file, name):
    zip_obj = ZipFile(name + '.zip', 'a')
    zip_obj.write(file, os.path.basename(file))
    zip_obj.close()


def delete_file(file):
    os.remove(file)


def validation_functions(wb):
    sheet = wb["Validation"]
    lists = "Validation"

    good = []
    bad = []

    not_a_validation = []
    bad_type = []
    false_operator = []
    wrong_formula = []

    validation_data = []
    for data_val in sheet.data_validations.dataValidation:
        cell_data = []
        types = data_val.type
        adress = data_val.sqref
        operator = data_val.operator
        formula1 = data_val.formula1
        formula2 = data_val.formula2
        cell_data.extend([adress, types, operator, formula1, formula2])
        validation_data.append(cell_data)

    for i in range(len(validation_data)):
        if (str(validation_data[i][0]).find("C3")) != -1:
            if str(validation_data[i][1]) == "whole" or str(validation_data[i][1]) == "decimal":
                if str(validation_data[i][2]) == "greaterThanOrEqual":
                    for j in range(5):
                        good.append("C" + str(j + 3))
                else:
                    for j in range(5):
                        bad.append("C" + str(j + 3))
                    wrong_formula.append("C3")
            else:
                for j in range(5):
                    bad.append("C" + str(j + 3))
                bad_type.append("C3")
        if (str(validation_data[i][0]).find("E3")) != -1:
            if str(validation_data[i][1]) == "date":
                if str(validation_data[i][3]) == "$H$1" and str(validation_data[i][4]) == "$H$2":
                    for j in range(5):
                        good.append("E" + str(j + 3))
                else:
                    for j in range(5):
                        bad.append("E" + str(j + 3))
                    wrong_formula.append("E3")
            else:
                for j in range(5):
                    bad.append("E" + str(j + 3))
                bad_type.append("E3")

        if (str(validation_data[i][0]).find("D3")) != -1:
            if str(validation_data[i][1]) == "custom":
                if str(validation_data[i][3]) == 'ISTEXT(D3)':
                    for j in range(5):
                        good.append("D" + str(j + 3))
                else:
                    for j in range(5):
                        bad.append("D" + str(j + 3))
                    wrong_formula.append("D3")
            else:
                for j in range(5):
                    bad.append("D" + str(j + 3))
                bad_type.append("D3")

        if (str(validation_data[i][0]).find("B3")) != -1:
            if str(validation_data[i][1]) == "list":
                if str(validation_data[i][3]) == '$G$5:$G$9' or str(validation_data[i][3]) == '$G$6:$G$9':
                    for j in range(5):
                        good.append("B" + str(j + 3))
                else:
                    for j in range(5):
                        bad.append("B" + str(j + 3))
                    wrong_formula.append("B3")
            else:
                for j in range(5):
                    bad.append("B" + str(j + 3))
                bad_type.append("B3")

    for i in good:
        cell_change_colour(wb, lists, i, "33FF33")

    check_list = ["C3", "E3", "D3", "B3"]
    for i in check_list:
        if i not in good:
            for j in range(5):
                bad.append(i[0] + str(j + 3))
            not_a_validation.append(i)

    for i in bad:
        cell_change_colour(wb, lists, i, "FF6666")

    for i in not_a_validation:
        vals = str(i[0]) + str(int(i[1:]) + 7)
        cell_write(wb, lists, vals, "Not a validation")
        cell_change_colour(wb, lists, vals, "FDDA0D")

    for i in bad_type:
        vals = str(i[0]) + str(int(i[1:]) + 7)
        cell_write(wb, lists, vals, "Wrong Type")
        cell_change_colour(wb, lists, vals, "FDDA0D")

    for i in false_operator:
        vals = str(i[0]) + str(int(i[1:]) + 7)
        cell_write(wb, lists, vals, "False operator")
        cell_change_colour(wb, lists, vals, "FDDA0D")

    for i in wrong_formula:
        vals = str(i[0]) + str(int(i[1:]) + 7)
        cell_write(wb, lists, vals, "Wrong formula")
        cell_change_colour(wb, lists, vals, "FDDA0D")


def conditional_function(wb):
    sheet = wb["Conditional formatting"]
    lists = "Conditional formatting"
    answer = []

    good = []
    bad = []

    not_a_validation = []

    for row in sheet.conditional_formatting:
        wz = str(row.cfRule[0]).split(", ")

        answer.append(
            [row.cells, str(str(wz[0])).replace("<openpyxl.formatting.rule.Rule object>\nParameters:\ntype=", " "),
             wz[4],
             wz[16]])

    for i in range(len(answer)):
        if (str(answer[i][0]).find("D2")) != -1:
            if str(answer[i][1]).find(" 'cellIs'") != -1:
                if str(answer[i][2]).find("greaterThan") != -1:
                    if str(answer[i][3]).find('200') != -1:
                        good.append("D26")
                    else:
                        bad.append("D26")
                else:
                    bad.append("D26")
            else:
                bad.append("D26")

        if (str(answer[i][0]).find("C2")) != -1:
            if str(answer[i][1]).find("containsText") != -1:
                if str(answer[i][2]).find("containsText") != -1:
                    if str(answer[i][3]).find('SEARCH') != -1:
                        good.append("C26")
                    else:
                        bad.append("C26")
                else:
                    bad.append("C26")
            else:
                bad.append("C26")

        if (str(answer[i][0]).find("E2")) != -1:
            if str(answer[i][1]).find("colorScale") != -1:
                if str(answer[i][3]).find('max') != -1:
                    good.append("E26")
                else:
                    bad.append("E26")

            else:
                bad.append("E26")

        if (str(answer[i][0]).find("H2")) != -1:
            if str(answer[i][1]).find("cellIs") != -1:
                if str(answer[i][2]).find("greaterThan") != -1:
                    if str(answer[i][3]).find('$J$2+15') != -1:
                        good.append("H26")
                    else:
                        bad.append("H26")
                else:
                    bad.append("H26")
            else:
                bad.append("H26")

    for i in good:
        cell_change_colour(wb, lists, i, "33FF33")
        cell_write(wb, lists, i, "Conditional format is correct")

    check_list = ["C26", "D26", "H26", "E26"]
    for i in check_list:
        if i not in good:
            bad.append(i)
            not_a_validation.append(i)

    for i in bad:
        cell_change_colour(wb, lists, i, "FF6666")
        cell_write(wb, lists, i, "Conditional format incorrect")


def file_check(file):
    if file.endswith("\n"):
        file = file[:-1]
    if file.endswith("false"):
        file = file[:-5]
    wb_start = r""
    file = wb_start + str(file)

    this_file = file.replace("\\", "/")
    copy_file(this_file)
    this_file = this_file.replace(".xlsx", "_copy.xlsx")
    return this_file


def finish():
    print("-------------------")
    print("Files are zipped")

    time.sleep(10)


def end_of_testing(wb, this_file, count, file_list, name):
    wb.save(this_file)
    wb.close()
    print(str(count) + "/" + str(len(file_list)) + " have been controlled")
    add_file_to_zip_without_directory(this_file, name)
    delete_file(this_file)


def script_start():
    print("Script started")
    print("Zip archive created")

